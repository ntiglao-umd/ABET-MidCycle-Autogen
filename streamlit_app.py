from __future__ import annotations

import re
import tempfile
from collections import Counter
from datetime import date, datetime, time
from pathlib import Path

import streamlit as st
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple

DEFAULT_OUTPUT_NAME = "ABET Mid-Cycle Autogen.xlsx"
BLUE = "#4F81BD"
COHORT_COLORS = [
    "#4F81BD",  # blue
    "#C0504D",  # red
    "#F2C811",  # yellow
    "#9BBB59",  # green
    "#8064A2",  # purple
    "#4BACC6",  # teal
]


# -----------------------------------------------------------------------------
# Format conversion helpers: openpyxl style -> XlsxWriter format
# -----------------------------------------------------------------------------
def rgb_from_openpyxl_color(color):
    """Return #RRGGBB from an openpyxl color object when possible."""
    if color is None:
        return None

    if getattr(color, "type", None) == "rgb" and color.rgb:
        rgb = color.rgb
        if len(rgb) == 8:  # ARGB -> RGB
            rgb = rgb[2:]
        return f"#{rgb}"

    return None


def border_style(style):
    """Map common openpyxl border styles to XlsxWriter numeric styles."""
    if not style:
        return None
    return {
        "thin": 1,
        "medium": 2,
        "dashed": 3,
        "dotted": 4,
        "thick": 5,
        "double": 6,
        "hair": 7,
        "mediumDashed": 8,
        "dashDot": 9,
        "mediumDashDot": 10,
        "dashDotDot": 11,
        "mediumDashDotDot": 12,
        "slantDashDot": 13,
    }.get(style, 1)


def format_from_cell(cell, out_wb, format_cache):
    """Create/reuse an XlsxWriter format that approximates the source cell style."""
    key = cell.style_id
    if key in format_cache:
        return format_cache[key]

    props = {}

    if cell.number_format:
        props["num_format"] = cell.number_format

    font = cell.font
    if font:
        if font.name:
            props["font_name"] = font.name
        if font.sz:
            props["font_size"] = font.sz
        if font.bold:
            props["bold"] = True
        if font.italic:
            props["italic"] = True
        if font.underline:
            props["underline"] = True
        if font.strike:
            props["font_strikeout"] = True
        color = rgb_from_openpyxl_color(font.color)
        if color:
            props["font_color"] = color

    fill = cell.fill
    if fill and fill.fill_type == "solid":
        fg = rgb_from_openpyxl_color(fill.fgColor)
        if fg:
            props["bg_color"] = fg
            props["pattern"] = 1

    alignment = cell.alignment
    if alignment:
        if alignment.horizontal:
            props["align"] = alignment.horizontal
        if alignment.vertical:
            props["valign"] = alignment.vertical
        if alignment.wrap_text:
            props["text_wrap"] = True
        if alignment.shrink_to_fit:
            props["shrink"] = True
        if alignment.text_rotation:
            props["rotation"] = alignment.text_rotation
        if alignment.indent:
            props["indent"] = int(alignment.indent)

    border = cell.border
    if border:
        for side_name, xlsx_name in {
            "left": "left",
            "right": "right",
            "top": "top",
            "bottom": "bottom",
        }.items():
            side = getattr(border, side_name)
            bstyle = border_style(side.style)
            border_style = _border_style(side.style if side is not None else None)
            if bstyle:
                props[xlsx_name] = bstyle
                color = rgb_from_openpyxl_color(side.color)
                if color:
                    props[f"{xlsx_name}_color"] = color

    fmt = out_wb.add_format(props)
    format_cache[key] = fmt
    return fmt

def _border_style(style):
    if not style:
        return None

    return {
        "thin": 1,
        "medium": 2,
        "thick": 5,
        "dashed": 3,
        "dotted": 4,
        "double": 6,
        "hair": 7,
        "mediumDashed": 8,
        "dashDot": 9,
        "mediumDashDot": 10,
        "dashDotDot": 11,
        "mediumDashDotDot": 12,
        "slantDashDot": 13,
    }.get(style, 1)

def write_cell(ws, row0, col0, cell, fmt):
    """Write a single cell while preserving formulas and date/time values."""
    value = cell.value

    if value is None:
        ws.write_blank(row0, col0, None, fmt)
    elif isinstance(value, str) and value.startswith("="):
        ws.write_formula(row0, col0, value, fmt)
    elif isinstance(value, (datetime, date, time)):
        ws.write_datetime(row0, col0, value, fmt)
    else:
        ws.write(row0, col0, value, fmt)


def detect_cohorts(src_ws):
    cohorts = []
    for r in range(1, src_ws.max_row + 1):
        val = src_ws.cell(r, 1).value
        if val and re.search(r"(Fall|Spring|Summer)\s+\d{4}", str(val), re.IGNORECASE):
            cohorts.append({"name": str(val).strip(), "start": r})

    for i in range(len(cohorts)):
        cohorts[i]["end"] = cohorts[i + 1]["start"] - 1 if i < len(cohorts) - 1 else src_ws.max_row

    return cohorts


def detect_dimension_columns(src_ws):
    dim_cols = []
    for r in range(1, min(15, src_ws.max_row) + 1):
        vals = [
            str(src_ws.cell(r, c).value).strip() if src_ws.cell(r, c).value is not None else ""
            for c in range(1, src_ws.max_column + 1)
        ]
        if "UID" in vals:
            for c, v in enumerate(vals, start=1):
                if v in ["Dim 1", "Dim 2"]:
                    dim_cols.append((v, c))
            break
    return dim_cols


def build_autogen_workbook(
    input_xlsx_path: str | Path,
    output_xlsx_path: str | Path,
    chart_width: int = 600,
    chart_height: int = 450,
    rounded_chart_corners: bool = True,
    chart_block_height: int = 25,
    chart_row_offset: int = 7,
):
    """Generate the ABET mid-cycle workbook from an uploaded Excel file."""
    src_wb = load_workbook(input_xlsx_path, data_only=False)
    out_wb = xlsxwriter.Workbook(output_xlsx_path)
    format_cache = {}

    fmt_percent = out_wb.add_format({"num_format": "0%", "font_name": "Calibri", "font_size": 11})
    fmt_header = out_wb.add_format({"font_name": "Calibri", "font_size": 11, "bold": True, "border": 1})
    fmt_cell = out_wb.add_format({"font_name": "Calibri", "font_size": 11, "border": 1})

    summary = []

    for src_ws in src_wb.worksheets:
        ws_name = src_ws.title[:31]
        ws = out_wb.add_worksheet(ws_name)
        max_row = src_ws.max_row
        max_col = src_ws.max_column

        if src_ws.sheet_view.showGridLines is False:
            ws.hide_gridlines(2)
        if src_ws.sheet_view.zoomScale:
            ws.set_zoom(src_ws.sheet_view.zoomScale)
        if src_ws.freeze_panes:
            fp = src_ws.freeze_panes
            if isinstance(fp, str):
                fr, fc = coordinate_to_tuple(fp)
                ws.freeze_panes(fr - 1, fc - 1)

        for r in range(1, max_row + 1):
            rd = src_ws.row_dimensions[r]
            if rd.height or rd.hidden:
                ws.set_row(r - 1, rd.height, None, {"hidden": rd.hidden})

        for c in range(1, max_col + 1):
            col_letter = src_ws.cell(1, c).column_letter
            cd = src_ws.column_dimensions[col_letter]
            width = cd.width if cd.width is not None else 8.43
            ws.set_column(c - 1, c - 1, width, None, {"hidden": cd.hidden})

        merged_ranges = {}
        merged_non_topleft = set()
        for merged in src_ws.merged_cells.ranges:
            min_col, min_row, max_col_m, max_row_m = merged.bounds
            merged_ranges[(min_row, min_col)] = (min_row, min_col, max_row_m, max_col_m)
            for rr in range(min_row, max_row_m + 1):
                for cc in range(min_col, max_col_m + 1):
                    if (rr, cc) != (min_row, min_col):
                        merged_non_topleft.add((rr, cc))

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                if (r, c) in merged_non_topleft:
                    continue

                cell = src_ws.cell(r, c)
                fmt = format_from_cell(cell, out_wb, format_cache)

                if (r, c) in merged_ranges:
                    min_row, min_col, max_row_m, max_col_m = merged_ranges[(r, c)]
                    ws.merge_range(
                        min_row - 1,
                        min_col - 1,
                        max_row_m - 1,
                        max_col_m - 1,
                        cell.value if cell.value is not None else "",
                        fmt,
                    )
                else:
                    write_cell(ws, r - 1, c - 1, cell, fmt)

        cohorts = detect_cohorts(src_ws)
        dim_cols = detect_dimension_columns(src_ws)
        summary.append({"Worksheet": ws_name, "Cohorts": len(cohorts), "Dimensions": ", ".join(d[0] for d in dim_cols)})

        if not cohorts or not dim_cols:
            continue

        chart_start_row = max_row + 5

        for dim_idx, (dim_name, dim_col) in enumerate(dim_cols):
            table_row = chart_start_row + dim_idx * chart_block_height
            levels = [
                (4, "4 - Mastery"),
                (3, "3 - Competent"),
                (2, "2 - Developing"),
                (1, "1 - Undeveloped"),
            ]

            ws.write(table_row - 1, 0, "Competency Level", fmt_header)
            for i, (_, label) in enumerate(levels, start=1):
                ws.write(table_row - 1 + i, 0, label, fmt_cell)

            for cohort_idx, cohort in enumerate(cohorts, start=1):
                col = cohort_idx
                ws.write(table_row - 1, col, cohort["name"], fmt_header)

                ratings = []
                for r in range(cohort["start"], cohort["end"] + 1):
                    val = src_ws.cell(r, dim_col).value
                    try:
                        if val is not None and str(val).strip() != "":
                            ratings.append(int(float(val)))
                    except Exception:
                        pass

                counts = Counter(ratings)
                total = len(ratings) if ratings else 1

                for i, (score, _) in enumerate(levels, start=1):
                    ws.write_number(table_row - 1 + i, col, counts.get(score, 0) / total, fmt_percent)

            chart = out_wb.add_chart({"type": "column"})
            chart.set_style(10)
            #chart.set_gap(80)
            # Set chart gap width; compatible with older/newer XlsxWriter versions
            if hasattr(chart, "set_gap"):
                chart.gap_width = 80
            else:
                chart.gap = 80
            chart.overlap = 0

            for cohort_idx, cohort in enumerate(cohorts, start=1):
                col = cohort_idx
                color = COHORT_COLORS[(cohort_idx - 1) % len(COHORT_COLORS)]
                series_opts = {
                    "name": [ws_name, table_row - 1, col],
                    "categories": [ws_name, table_row, 0, table_row + 3, 0],
                    "values": [ws_name, table_row, col, table_row + 3, col],
                    "fill": {"color": color},
                    "border": {"color": "#000000"},
                }
                if len(cohorts) == 1:
                    series_opts["fill"] = {"color": BLUE}
                chart.add_series(series_opts)

            chart.set_title({
                "name": f"{dim_name} Analysis",
                "name_font": {"bold": True, "size": 18, "color": "#666666"},
            })
            chart.set_x_axis({
                "name": " ",
                "name_font": {"bold": True},
                "label_position": "low",
                "line": {"color": "#808080"},
            })
            chart.set_y_axis({
                "name": " ",
                "name_font": {"bold": True},
                "num_format": "0.00%",
                "min": 0,
                "max": 1.0,
                "major_unit": 0.1,
                "line": {"color": "#808080"},
                "major_gridlines": {
                    "visible": True,
                    "line": {"color": "#BFBFBF", "width": 0.5},
                },
            })
            chart.set_legend({"position": "right"})
            chart.set_size({"width": chart_width, "height": chart_height})
            chart.set_chartarea({
                "border": {"color": "#808080"},
                "fill": {"color": "white"},
                "rounded_corners": rounded_chart_corners,
            })
            chart.set_plotarea({
                "fill": {"color": "white"},
                "border": {"none": True},
            })

            ws.insert_chart(table_row + chart_row_offset, 7, chart, {"x_offset": 10, "y_offset": 10})

    out_wb.close()
    return summary


# -----------------------------------------------------------------------------
# Streamlit UI
# -----------------------------------------------------------------------------
st.set_page_config(page_title="ABET Mid-Cycle Autogen", page_icon="📊", layout="centered")

st.title("ABET Mid-Cycle Autogen")
st.write("Upload the ABET Student Assignment Sheet Excel file to generate a formatted workbook with Dim 1 and Dim 2 analysis charts.")

uploaded_file = st.file_uploader("Upload Excel workbook", type=["xlsx"])

with st.expander("Chart options", expanded=False):
    chart_width = st.number_input("Chart width", min_value=400, max_value=1200, value=600, step=10)
    chart_height = st.number_input("Chart height", min_value=300, max_value=900, value=450, step=10)
    rounded_chart_corners = st.checkbox("Rounded chart frame corners", value=True)
    chart_block_height = st.number_input("Rows reserved per chart block", min_value=20, max_value=60, value=25, step=1)
    chart_row_offset = st.number_input("Chart row offset below helper table", min_value=3, max_value=20, value=7, step=1)

output_name = st.text_input("Output filename", value=DEFAULT_OUTPUT_NAME)

if uploaded_file is not None:
    st.success(f"Uploaded: {uploaded_file.name}")

    if st.button("Generate workbook", type="primary"):
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                tmpdir_path = Path(tmpdir)
                input_path = tmpdir_path / "input.xlsx"
                output_path = tmpdir_path / output_name

                input_path.write_bytes(uploaded_file.getvalue())

                with st.spinner("Generating workbook..."):
                    summary = build_autogen_workbook(
                        input_xlsx_path=input_path,
                        output_xlsx_path=output_path,
                        chart_width=int(chart_width),
                        chart_height=int(chart_height),
                        rounded_chart_corners=rounded_chart_corners,
                        chart_block_height=int(chart_block_height),
                        chart_row_offset=int(chart_row_offset),
                    )

                output_bytes = output_path.read_bytes()

            st.success("Workbook generated successfully.")
            st.dataframe(summary, use_container_width=True)
            st.download_button(
                label="Download generated workbook",
                data=output_bytes,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as exc:
            st.error("The workbook could not be generated.")
            st.exception(exc)
else:
    st.info("Upload an .xlsx file to begin.")

st.caption("Run locally with: streamlit run ABET_mid_cycle_streamlit_app.py")
